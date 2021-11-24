import openpyxl
from pprint import pprint
from io import BytesIO


def excel_to_list_with_openpyxl(filename, in_memory=None):
    '''
    Lê planilha Excel (xlsx) e retorna uma lista de dicionários.
    Usando openpyxl
    '''
    if in_memory:
        # filename=BytesIO(input_excel.read())
        pass
    else:
        wb = openpyxl.load_workbook(filename)

    ws = wb.worksheets[0]  # read first sheet.

    first_line = []
    for col in ws[1]:
        first_line.append(col.value.strip())

    data = []
    for row in range(2, ws.max_row + 1):
        obj_dict = {}
        for col in range(1, ws.max_column + 1):
            obj_dict[first_line[col - 1]] = ws.cell(row=row, column=col).value
        data.append(obj_dict)
    return data


if __name__ == '__main__':
    filename = 'tabelaRossHeideck.xlsx'
    data = excel_to_list_with_openpyxl(filename)
    for item in data:
        pprint(item)
        print()
        print('IDADE EM % DE VIDA', item['IDADE EM % DE VIDA'])
        print('A', item['A'])
        print('B', item['B'])
        print('C', item['C'])
        print('H', item['H'])
        print('---------------')