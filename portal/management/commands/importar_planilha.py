'''
Usage:
    python manage.py importar_planilha --filename tabelaRossHeideck.xlsx
'''
import openpyxl
from django.core.management.base import BaseCommand
from portal.models import Tabelarossheideck


def excel_to_list_with_openpyxl(filename, in_memory=None):
    '''
    Lê planilha Excel (xlsx) e retorna uma lista de dicionários.
    Usando openpyxl
    '''
    if in_memory:
        # filename=BytesIO(input_excel.read())
        pass
    else:
        wb = openpyxl.load_workbook(filename)

    ws = wb.worksheets[0]  # read first sheet.

    first_line = []
    for col in ws[1]:
        first_line.append(col.value.strip())

    data = []
    for row in range(2, ws.max_row + 1):
        obj_dict = {}
        for col in range(1, ws.max_column + 1):
            obj_dict[first_line[col - 1]] = ws.cell(row=row, column=col).value
        data.append(obj_dict)
    return data


class Command(BaseCommand):
    help = 'Importa dados de uma planilha xlsx.'

    def add_arguments(self, parser):
        # Argumento nomeado (opcional)
        parser.add_argument(
            '--filename',
            help='Nome do arquivo.'
        )

    def handle(self, *args, **options):
        if options['filename']:
            data = excel_to_list_with_openpyxl(options['filename'])
            for item in data:
                Tabelarossheideck.objects.create(
                    idade_em_vida=item['IDADE EM % DE VIDA'],
                    a=item['A'],
                    b=item['B'],
                    c=item['C'],
                    d=item['D'],
                    e=item['E'],
                    f=item['F'],
                    g=item['G'],
                    h=item['H'],
                )

            self.stdout.write('Dados importados com sucesso.')
